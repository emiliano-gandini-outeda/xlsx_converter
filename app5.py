import pandas as pd
import openpyxl
import xlrd
import os
from datetime import datetime

# Detalles de Articulos Activos - Rediseñado

def get_column_letter(col_num):
    """Convierte número de columna a letra de Excel"""
    letters = []
    for i in range(1, 35):  # Hasta columna AH
        if i <= 26:
            letters.append(chr(64 + i))  # A-Z
        else:
            letters.append('A' + chr(64 + i - 26))  # AA-AH
    return letters[col_num - 1] if col_num <= len(letters) else f"Col{col_num}"

def clean_value(value, data_type='string'):
    """Limpia y convierte valores según el tipo de dato especificado"""
    if pd.isna(value) or value == '' or str(value).strip() == '{Sin Definir}' or str(value).strip() == '':
        return "Dato no Definido"
    
    if data_type == 'integer':
        try:
            if isinstance(value, str):
                value = value.replace(',', '').replace(' ', '')
            return int(float(value))
        except (ValueError, TypeError):
            return "Dato no Definido"
    elif data_type == 'float':
        try:
            if isinstance(value, str):
                value = value.replace(',', '').replace(' ', '')
            return float(value)
        except (ValueError, TypeError):
            return "Dato no Definido"
    elif data_type == 'importado':
        # Tipo especial para la columna importado que debe ser "Si" o "No"
        str_value = str(value).strip().lower()
        if str_value in ['si', 'sí', 'yes', 'y', '1', 'true', 'verdadero']:
            return "Si"
        elif str_value in ['no', 'n', '0', 'false', 'falso']:
            return "No"
        else:
            return "No"  # Valor por defecto
    else:  # string
        return str(value).strip() if str(value).strip() != '' else "Dato no Definido"

def process_file(file_path):
    """Procesa el archivo de inventario y extrae la información de proveedores y artículos"""
    
    try:
        # Determinar el tipo de archivo y cargar apropiadamente
        file_extension = os.path.splitext(file_path)[1].lower()
        
        if file_extension == '.xlsx':
            # Para archivos .xlsx usar openpyxl
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            max_row = sheet.max_row
            is_xlsx = True
        elif file_extension == '.xls':
            # Para archivos .xls usar xlrd
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            max_row = sheet.nrows
            is_xlsx = False
        else:
            raise ValueError("Formato de archivo no soportado. Solo se admiten .xlsx y .xls")
        
        print(f"Procesando archivo {file_extension.upper()}: {os.path.basename(file_path)}")
        
        # Lista para almacenar todos los datos procesados
        processed_data = []
        
        # Función auxiliar para obtener valor de celda según el tipo de archivo
        def get_cell_value(row, col):
            if is_xlsx:
                # Para openpyxl (xlsx) - usar notación de letra
                col_letter = get_column_letter(col)
                cell = sheet[f'{col_letter}{row}']
                return cell.value
            else:
                # Para xlrd (xls) - usar índices (row-1, col-1 porque xlrd usa base 0)
                try:
                    return sheet.cell_value(row - 1, col - 1)
                except IndexError:
                    return None
        
        # Recorrer todas las filas para buscar "Proveedor:"
        current_row = 1
        proveedores_encontrados = 0
        
        print(f"Iniciando búsqueda en {max_row} filas...")
        
        while current_row <= max_row:
            # Buscar "Proveedor:" en la columna B (columna 2)
            cell_b_value = get_cell_value(current_row, 2)
            
            if cell_b_value and 'Proveedor:' in str(cell_b_value):
                proveedores_encontrados += 1
                print(f"Encontrado proveedor #{proveedores_encontrados} en fila {current_row}")
                
                # Leer información del proveedor de la misma fila
                # Columna F (6): ID del proveedor (Integer)
                id_proveedor = clean_value(get_cell_value(current_row, 6), 'integer')
                # Columna M (13): Nombre del proveedor (String)
                nombre_proveedor = clean_value(get_cell_value(current_row, 13), 'string')
                
                print(f"ID Proveedor: {id_proveedor}, Nombre: {nombre_proveedor}")
                
                # Pasar a la siguiente fila para empezar a leer artículos
                current_row += 1
                
                # Leer artículos hasta encontrar otro proveedor o fin de archivo
                while current_row <= max_row:
                    # Verificar si en la columna B hay otro "Proveedor:" (fin de artículos de este proveedor)
                    cell_b_next_value = get_cell_value(current_row, 2)
                    if cell_b_next_value and 'Proveedor:' in str(cell_b_next_value):
                        # Encontramos otro proveedor, salir del bucle de artículos
                        break
                    
                    # Leer información del artículo según las especificaciones:
                    # Columna B (2): ID del Articulo (String)
                    id_articulo = clean_value(get_cell_value(current_row, 2), 'string')
                    # Columna I (9): Nombre del articulo (string)
                    nombre_articulo = clean_value(get_cell_value(current_row, 9), 'string')
                    # Columna S (19): Stock Minimo (float)
                    stock_minimo = clean_value(get_cell_value(current_row, 19), 'float')
                    # Columna V (22): Estado del Producto (String)
                    estado_producto = clean_value(get_cell_value(current_row, 22), 'string')
                    # Columna Z (26): Importado (string "Si" o "No")
                    importado = clean_value(get_cell_value(current_row, 26), 'importado')
                    # Columna AC (29): Codigo para proveedor (string)
                    codigo_proveedor = clean_value(get_cell_value(current_row, 29), 'string')
                    
                    # Debug: mostrar valores de la fila actual
                    print(f"Fila {current_row}: ID='{id_articulo}', Nombre='{nombre_articulo}', Stock='{stock_minimo}'")
                    
                    # Solo agregar si hay información válida del artículo
                    # Verificar que al menos tenga ID o nombre del artículo
                    if (id_articulo != "Dato no Definido" or 
                        nombre_articulo != "Dato no Definido"):
                        
                        print(f"Artículo válido encontrado - ID: {id_articulo}, Nombre: {nombre_articulo}")
                        
                        articulo_data = {
                            'ID Proveedor': id_proveedor,
                            'Nombre Proveedor': nombre_proveedor,
                            'ID Articulo': id_articulo,
                            'Nombre Articulo': nombre_articulo,
                            'Stock Minimo': stock_minimo,
                            'Estado del Producto': estado_producto,
                            'Importado': importado,
                            'Codigo para Proveedor': codigo_proveedor
                        }
                        processed_data.append(articulo_data)
                    
                    current_row += 1
            else:
                current_row += 1
        
        if not processed_data:
            raise ValueError("No se encontraron proveedores con el formato esperado en el archivo")
        
        print(f"\nResumen del procesamiento:")
        print(f"- Proveedores encontrados: {proveedores_encontrados}")
        print(f"- Total de artículos procesados: {len(processed_data)}")
        
        # Crear DataFrame con los datos procesados
        df = pd.DataFrame(processed_data)
        
        # Reordenar columnas según especificación
        column_order = [
            'ID Proveedor',
            'Nombre Proveedor', 
            'ID Articulo',
            'Nombre Articulo',
            'Stock Minimo',
            'Estado del Producto',
            'Importado',
            'Codigo para Proveedor'
        ]
        df = df[column_order]
        
        # Generar nombre de archivo de salida
        base_filename = os.path.splitext(os.path.basename(file_path))[0]
        output_filename = f"{base_filename}_PROCESADO.xlsx"
        output_path = os.path.join(os.path.dirname(file_path), output_filename)
        
        # Guardar archivo Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Inventario Procesado', index=False)
            
            # Obtener la hoja y aplicar formato
            worksheet = writer.sheets['Inventario Procesado']
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"Archivo procesado guardado en: {output_path}")
        print(f"Total de artículos procesados: {len(df)}")
        
        # Mostrar preview de los primeros registros
        print("\nPreview de los primeros 3 registros:")
        print(df.head(3).to_string())
        
        return output_path
        
    except Exception as e:
        print(f"Error procesando archivo: {str(e)}")
        raise Exception(f"Error al procesar el archivo de inventario: {str(e)}")

if __name__ == "__main__":
    # Para pruebas locales
    test_files = ["test_inventario.xlsx", "test_inventario.xls"]
    for test_file in test_files:
        if os.path.exists(test_file):
            print(f"\nProcesando: {test_file}")
            try:
                result = process_file(test_file)
                print(f"Archivo procesado: {result}")
            except Exception as e:
                print(f"Error procesando {test_file}: {e}")
        else:
            print(f"Archivo de prueba no encontrado: {test_file}")